import sys
import psycopg2
import clipboard as clip
from openpyxl import Workbook
from openpyxl.styles import Font
from querys import *
from PyQt5 import *

from PyQt5.QtWidgets import QMainWindow, QApplication,QLineEdit
from PyQt5.uic import loadUi 
from PyQt5.QtCore import Qt
from datetime import date
from PyQt5 import QtGui


global rango, numpedido, pedidos, disponible,subpedidos,INICIAL,FINAL
import os
from mensaje import *

pedidos=[]
subpedidos=[]



class VentanaPrincipal(QMainWindow):
	def __init__(self):
		super(VentanaPrincipal, self).__init__()
		loadUi('main.ui', self)
		
		self.frame_detallepedido.hide()
		
		#CAMBIAR DE PAGINAS
		self.btn_nuevopedido.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_nuevopedido))#cambia de pagina
		self.btn_nuevopedido.clicked.connect(lambda: self.signal_barra.setText("PEDIDOS"))

		self.btn_nuevosubpedido.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_nuevosubpedido)) #cambia de pagina
		self.btn_nuevosubpedido.clicked.connect(lambda: self.signal_barra.setText("SUB PEDIDOS"))

		self.btn_listar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_listar)) #cambia de pagina
		self.btn_listar.clicked.connect(lambda: self.signal_barra.setText("LISTAR"))
		
		self.btn_rendicion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_rendicion)) #cambia de pagina
		self.btn_rendicion.clicked.connect(lambda: self.signal_barra.setText("RENDICION"))

		self.btn_configuracion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_config)) #cambia de pagina
		self.btn_configuracion.clicked.connect(lambda: self.signal_barra.setText("CONFIGURACION"))

		self.btn_deposito.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_deposito)) #cambia de pagina
		self.btn_deposito.clicked.connect(lambda: self.signal_barra.setText("LOCKERS"))

		self.btn_rotulos.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_rotulos)) #cambia de pagina
		self.btn_rotulos.clicked.connect(lambda: self.signal_barra.setText("ROTULOS"))

		self.btn_nuevagestion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_gestion)) #cambia de pagina
		self.btn_nuevagestion.clicked.connect(lambda: self.signal_barra.setText("GESTIONES"))

		self.btn_envios.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_envios))
		self.btn_envios.clicked.connect(lambda: self.signal_barra.setText("ENVIOS"))

		self.btn_estampillas.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_estampillas))
		self.btn_estampillas.clicked.connect(lambda: self.signal_barra.setText("ESTAMPILLAS"))

		self.btn_nuevopedido.clicked.connect(self.iniciarpedido)
		
		#FUNCION DE LOS BOTONES
		#PAGINA PEDIDOS
		self.btn_ingresarpedido.clicked.connect(self.NuevoPedido)
		self.btn_ingresarpedido.clicked.connect(self.limpiar)
		
		self.btn_printdetalle.clicked.connect(self.imprimirticket)
		
		self.rb_seriea.clicked.connect(self.iniciarpedido)
		self.rb_serieb.clicked.connect(self.iniciarpedido)
		self.rb_estampillas.clicked.connect(self.refresh_estampillas)
		self.rb_general.toggled.connect(self.toggle_frame)
		
		
			
		#boton salir
		self.btn_salir.clicked.connect(salir)
		
		#pagina subpedidos
		self.btn_verpedidos.clicked.connect(self.verpedidos)
		self.btn_subvalidar.clicked.connect(self.validarpedido)
		self.btn_subingresar.clicked.connect(self.nuevosubpedido)
		self.btn_subingresar.clicked.connect(self.limpiar)
		self.combo_asociados_subpedidos.activated.connect(self.traeregistrosp)
		self.combo_asociados_subpedidos.activated.connect(self.verpedidos)
		self.tb_verpedidos.itemDoubleClicked.connect(self.completanumpedido)
		self.btn_copy_inicio.clicked.connect(self.clipinicio)
		self.btn_refresh.clicked.connect(self.refresh_pedidos)
		self.btn_deshacer.clicked.connect(self.deshacerSubpedido)
		self.btn_eliminarSubpedido.clicked.connect(self.eliminar_subpedido)
		
		
		#pagina listar
		self.btn_listarlistar.clicked.connect(self.listar)
		self.combo_asociados_listar.activated.connect(self.traeregistrolistar)
		self.btn_pedidoexcell.clicked.connect(self.listartoexcell)
		
		#PAGINA RENDICION
		self.btn_consultar.clicked.connect(self.rendir)
		self.btn_rendir.clicked.connect(self.exportarrendicion)
		
		#PAGINA CONFIGURACION
		self.btn_definir_locker.clicked.connect(self.setearlockers)
		self.btn_cargarasociado.clicked.connect(self.altasocio)
		self.btn_agregarrango.clicked.connect(self.agregarrango)
		self.btn_ver_rangos.clicked.connect(self.verrangos)
		self.btn_definir_rango.clicked.connect(self.setearrango)
		self.tb_rangos.itemDoubleClicked.connect(self.rangoselected)
		self.btn_terminar_rango.clicked.connect(self.bajaderango)
		self.btn_componer_rango.clicked.connect(self.componer_rango)
		self.btn_traer_rangos.clicked.connect(self.traeRango)
		self.btn_traer_rangos_pedido.clicked.connect(self.recupera_rango_pedido)
		self.btn_corregir_pedido.clicked.connect(self.componer_pedido)
		
		#PROPIEDADES
		self.txt_indice_rotulos.hide()
		self.frm_rangosExternos.hide()
		self.rb_porrncyfs.setChecked(True)
		self.rb_todos.setChecked(True)
		self.fechadesde_rendicion.setDate(date.today())
		self.fechahasta_rendicion.setDate(date.today())
		self.fecha_desde_rotulos.setDate(date.today())
		self.fecha_hasta_rotulos.setDate(date.today())
		self.fecha_desde_listar.setDate(date.today())
		self.fecha_hasta_listar.setDate(date.today())
		self.fecha_desde_envios.setDate(date.today())
		self.fecha_hasta_envios.setDate(date.today())
		self.fecha_desde_estampillas.setDate(date.today())
		self.fecha_hasta_estampillas.setDate(date.today())
		self.signal_gestion_indice.hide()
		self.cbx_porfecha.stateChanged.connect(lambda:self.fecha_desde_listar.setEnabled(True))
		self.cbx_porfecha.stateChanged.connect(lambda:self.fecha_hasta_listar.setEnabled(True))
		
		#AJUSTAR CONTENIDO DE LAS CELDAS
		headertb_gestiones = self.tb_gestiones.horizontalHeader()
		headertb_gestiones.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
		
		
		self.tb_verpedidos.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)

		
		headertb_lockers = self.tb_lockers.horizontalHeader()
		headertb_lockers.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
		
		headertb_rangos = self.tb_rangos.horizontalHeader()
		headertb_rangos.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
		
		headertb_rendicion = self.tb_rendicion.horizontalHeader()
		headertb_rendicion.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
		
		headertb_asociados_envios = self.tb_asociados_envios.horizontalHeader()
		headertb_asociados_envios.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
						
		headertb_asociados_pedidos2 = self.tb_asociados_pedidos2.horizontalHeader()
		headertb_asociados_pedidos2.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
		
		headertb_pedidos_nuevopedido = self.tb_pedidos_nuevopedido.horizontalHeader()
		headertb_pedidos_nuevopedido.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)

		headertb_estampillas = self.tb_estampillas.horizontalHeader()
		headertb_estampillas.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)

		headertb_envios = self.tb_envioscreados.horizontalHeader()
		headertb_envios.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)

		
		
		
		#PAGINA DEPOSITO
		self.btn_almacenar.clicked.connect(self.almacenar)
		self.tb_lockers.itemDoubleClicked.connect(self.lockerselected)
		self.btn_despachar.clicked.connect(self.despachar)
		self.btn_buscar_locker.clicked.connect(self.filtrarlockers)
		
		#PAGINA ROTULOS
		self.btn_rotulos.clicked.connect(self.listarimpresiones)
		self.btn_rotulos.clicked.connect(self.traerstock)
		self.btn_ingresar_rotulos.clicked.connect(self.nuevaimpresion)
		self.cb_razonsocial_rotulos.activated.connect(self.traeregistrorotulos)
		self.btn_listarimpresiones.clicked.connect(self.listarimpresiones)
		self.btn_definir_rotulos.clicked.connect(self.cambiarestadorotulo)
		self.tb_rotulos.itemDoubleClicked.connect(self.impresionselected)
		self.btn_eliminar_impresion.clicked.connect(self.eliminarimpresion)
		self.btn_ingresar_stock.clicked.connect(self.cargaStock)
		self.btn_modificar_rotulos.clicked.connect(self.editarcantidadrotulo)
		self.btn_limpiar_busqueda_rotulos.clicked.connect(self.limpiar_formulario_busqueda_impresiones)
		self.btn_modificar_tipo.clicked.connect(self.modificar_tipo)
		self.btn_exportar_impresiones.clicked.connect(self.exportar_listado_rotulos)
		
		#PAGINA GESTIONES
		self.btn_agregar_nueva.clicked.connect(self.nuevagestion)
		self.btn_buscar_gestiones.clicked.connect(self.listargestiones)
		self.tb_gestiones.itemDoubleClicked.connect(self.gestionselected)
		self.btn_actualizargestion.clicked.connect(self.gestionupdate)
		self.btn_nuevagestion.clicked.connect(self.listargestiones)
		self.btn_buscar_gestiones_filtro.clicked.connect(self.filtrargestiones)
		self.btn_ingresarnota.clicked.connect(self.notaupdate)
		self.btn_eliminar_gestion.clicked.connect(self.eliminargestion)
		
		#PAGINA ENVIOS
		self.tb_asociados_envios.itemDoubleClicked.connect(self.asociado_selected)
		
		self.tb_subpedidos_envios.itemDoubleClicked.connect(self.pedido_envios_selected)
		self.btn_buscar_asociados.clicked.connect(self.filtrar_asociados)
		
		self.tb_asociados_pedidos2.itemDoubleClicked.connect(self.asociado_selected_pedidos)
		self.btn_buscarasociado_pedidos.clicked.connect(self.filtrar_asociados_pedidos)
		self.btn_crearenvio.clicked.connect(self.nuevoEnvio)
		self.btn_refresh_envios.clicked.connect(self.traerenvios)
		self.tb_envioscreados.itemDoubleClicked.connect(self.envio_selected)
		self.rb_envios_iqr.clicked.connect(self.asociado_selected)
		
		self.rb_envios_estampillas.clicked.connect(self.asociado_selected)
		self.rb_envios_anexo.clicked.connect(self.asociado_selected)
		self.btn_agregar_guia.clicked.connect(self.agregar_guia)
		self.btn_calcular_costo_envio.clicked.connect(self.calcular_costo_sugerido)
		self.btn_facturar_envio.clicked.connect(self.definir_envio_facturado)
		self.btn_agregar_fct.clicked.connect(self.agregar_factura)
		self.btn_eliminar_envio.clicked.connect(self.eliminar_envio)
		self.rb_xasociado_envios.clicked.connect(self.traerenvios)
		self.rb_todoslosenvios_envios.clicked.connect(self.traerenvios)
		
		
		#PAGINA ESTAMPILLAS
		self.combo_asociados_estampillas.activated.connect(self.traeregistroestampillas)
		self.btn_ingresar_estampillas.clicked.connect(self.asigna_estampillas)
		self.rb_estampillas.clicked.connect(self.refresh_estampillas)
		self.rb_anexo.clicked.connect(self.refresh_estampillas)
		self.btn_eliminarLinea.clicked.connect(self.eliminar_linea)
		self.btn_buscar_estampillas.clicked.connect(self.refresh_estampillas)

		
	
		
	def llenarcombo(self):
		
		asociados = q.traerasociados() #la consulta devuelve una tupla, por lo tanto hay que convertirla a str para llenar el combobox, se usa el metodo "".join()
		
		k=0
		for i in asociados:
			
			self.combo_asociados_subpedidos.addItem("".join(asociados[k]))
			self.combo_asociados_listar.addItem("".join(asociados[k]))
			self.cb_razonsocial_rotulos.addItem("".join(asociados[k]))
			self.cb_razonsocial.addItem("".join(asociados[k]))
			self.combo_asociados_gestiones.addItem("".join(asociados[k]))
			self.combo_asociados_estampillas.addItem("".join(asociados[k]))
			k=k+1
		

	def toggle_frame(self):
			# Mostrar u ocultar el frame basado en el estado del radio button
			if self.rb_general.isChecked():
				self.frm_rangosExternos.show()
			else:
				self.frm_rangosExternos.hide()



		
	def iniciarpedido(self):
		global INICIAL
		global FINAL
		global Numpedido
		global indice
		
		dato= q.traeultimopedido()
		
		if self.rb_seriea.isChecked():
			indice=1
		elif self.rb_serieb.isChecked():
			indice=2
		elif self.rb_general.isChecked():
			indice=3
				
		rangogeneral = q.recuperarango(indice)
		
		Numpedido=dato[0]
		INICIAL=int(rangogeneral[0])
		FINAL=int(rangogeneral[1])
		
		
		disponible = int(FINAL-INICIAL+1)
		self.signal_stock.setText(str(disponible))
		self.signal_inicial.setText(str(INICIAL))
		self.signal_final.setText(str(FINAL))
		
	def refresh_estampillas(self):
		global ESTAMPILLA_INICIAL
		global ESTAMPILLA_FINAL

		if self.rb_estampillas.isChecked():
			indice =4
			self.txt_dav.setText("")
			self.txt_variedad_estampillas.setText("")
			
			self.txt_envase_estampillas.setText("")
			



			self.txt_dav.setEnabled(True)
			self.txt_variedad_estampillas.setEnabled(True)			
			self.txt_categoria_estampillas.setEnabled(True)
			self.txt_envase_estampillas.setEnabled(True)

		else:
			indice =5
			self.txt_dav.setText("No se informa")
			self.txt_variedad_estampillas.setText("No se informa")
			
			self.txt_envase_estampillas.setText("No se informa")
			
							
			self.txt_dav.setEnabled(False)
			self.txt_variedad_estampillas.setEnabled(False)			
			self.txt_categoria_estampillas.setEnabled(False)
			self.txt_envase_estampillas.setEnabled(False)

		rango_recuperado =q.recuperarango(indice)
		self.signal_estampillas.setText(str(rango_recuperado[0]))
		self.signal_estampillas_final.setText(str(rango_recuperado[1]))
		ESTAMPILLA_INICIAL=int(rango_recuperado[0])
		ESTAMPILLA_FINAL=int(rango_recuperado[1])

		self.ver_estampillas()

	def exportar_informe_estampillas(self):
		pass
		
		
	
			
		
	def traeregistro(self):
		nombre = str(self.combo_asociados.currentText())
		
		
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_rncyfs.setText("".join(recuperado))
		
		
	def traeregistrosp(self):
		nombre = str(self.combo_asociados_subpedidos.currentText())
		
		
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_verpedido.setText("".join(recuperado))
	def traeregistrolistar(self):
		nombre = str(self.combo_asociados_listar.currentText())
		
		
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_listar.setText("".join(recuperado))
	
	def traeregistrorotulos(self):

		try:
			nombre = str(self.cb_razonsocial_rotulos.currentText())
			
			
			recuperado=q.getrncyfs(nombre)
			
				
			self.txt_rncyfs_rotulos.setText("".join(recuperado))
		except Exception as e:
			self.txt_rncyfs_rotulos.setText("")
			
			

	def traeregistroestampillas(self):


		nombre = str(self.combo_asociados_estampillas.currentText())
		
		
		recuperado=q.getrncyfs(nombre)

		if recuperado:
			
			self.signal_rncyfs_estampillas.setText("".join(recuperado))
			self.signal_razonsocial_estampillas.setText(nombre)

		else:
			self.signal_razonsocial_estampillas.setText("Seleccione Asociado...")
			self.signal_rncyfs_estampillas.setText("")

	
	def traeregistrogestiones(self):
		nombre = str(self.combo_asociados_gestiones.currentText())
		
		
		recuperado=q.getrncyfs(nombre)
		
		
		
			
		return ("".join(recuperado))
		
	def traepedidos(self):
		
		tablapedidos=q.getpedidos()
		totalfilas=len(tablapedidos)
		self.tb_pedidos_nuevopedido.setRowCount(totalfilas)
		
		fila=0
		
		
		for i in tablapedidos:
			
						
			self.tb_pedidos_nuevopedido.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_pedidos_nuevopedido.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_pedidos_nuevopedido.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_pedidos_nuevopedido.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3]))) 
			self.tb_pedidos_nuevopedido.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4]))) 
			self.tb_pedidos_nuevopedido.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5]))) 
			
			
			fila=fila+1
		
		
		
	def NuevoPedido(self):
		global INICIAL
		global FINAL
		global Numpedido
		global indice
		
		
		if self.txt_cantidad.text(): 
				
			
				
						
			
			disponible = FINAL-INICIAL+1
			
			cantidad = int(self.txt_cantidad.text())
			registro = str(self.txt_rncyfs.text())
			
			if cantidad <= disponible:

				P = Pedido(cantidad,registro,Numpedido) #crea nuevo pedido
				P.asignar(INICIAL) #envia como parametro el rango inicial

				
				
				P.showrango()
				pedidos.append(P)
				Numpedido= Numpedido+1
					
				estado="SIN USAR"
					
					
					
				fechapedido=str(date.today())
					
					
					#fechapedido=formatearfecha(fecha)
				flag =False	
				if self.rb_seriea.isChecked():
					serie="A"
					
				elif self.rb_serieb.isChecked():
					serie="B"
					
				elif self.rb_general.isChecked():
					serie="Grl."
					flag=True
					
					
					
				if flag is True:
					iniciogeneral = int(self.txt_inicioGral.text())
					finalgeneral = int(self.txt_finGral.text())
					q.cargapedido(Numpedido,registro,cantidad,iniciogeneral,finalgeneral,iniciogeneral,finalgeneral,estado,fechapedido,serie)
					
					
					
				else:
						
					q.cargapedido(Numpedido,registro,cantidad,INICIAL,INICIAL+cantidad-1,INICIAL,INICIAL+cantidad-1,estado,fechapedido,serie)
					
								
					self.frame_detallepedido.show()			
					self.signal_cantidad.setText(str(cantidad))
					self.signal_rncyfs.setText(str(registro))
					self.signal_inicio.setText(str(INICIAL))
					self.signal_fin.setText(str(INICIAL+cantidad-1))
					self.signal_numpedido.setText(str(Numpedido))
					
					#GUARDA DATOS PARA IMRIMIR
					nombre=q.traenombre(registro)
						
					ticket= open("ticket.txt","w")
						
					ticket.write("DETALLE DE PEDIDO\n")
					ticket.write("-----------------------------\n")
					ticket.write("RAZON SOCIAL: "+str("".join(nombre))+"\n")
					ticket.write("PEDIDO: "+str(Numpedido)+"\n")
					ticket.write("-----------------------------\n")
					ticket.write("RNCyFS: "+str(registro)+"\n")
					ticket.write("Fecha: "+str(fechapedido)+"\n")			
					ticket.write("Rango: ")
					ticket.write(str(INICIAL)+" - "+str(INICIAL+cantidad-1)+"\n")
					ticket.write("Serie: "+str(serie)+"\n")
					ticket.write("Cantidad: "+str(cantidad)+"\n")
						
						
					ticket.write("-----------------------------\n")
					ticket.close()
					INICIAL=INICIAL+cantidad
					q.actualizarangoenbd(INICIAL,FINAL,indice)
								
					
				c.cartel("AVISO","RANGO ASIGNADO CORRECTAMENTE",1)	
				q.incrementanpedido(Numpedido)	
				
				
				self.iniciarpedido()
				self.traepedidos()
				tipo="Rotulo IQR"
				estado="4-Esperando Nota por KG"
				
				q.altagestion(registro,tipo,estado,fechapedido,cantidad)
				
				
			else:
				
				c.cartel("ERROR","NO HAY STOCK SUFICIENTE",3)
					
					
		else:
			
			c.cartel("ERROR","INGRESE CANTIDAD",3)


	def asigna_estampillas(self):
		global ESTAMPILLA_INICIAL
		global ESTAMPILLA_FINAL

		try:

			if self.signal_rncyfs_estampillas.text():

				rncyfs=str(self.signal_rncyfs_estampillas.text())
				
				especie=str(self.txt_estampillas_especie.currentText())
				
				campana = int(self.txt_campana_estampillas.currentText())
				cantidad =int(self.txt_cantidad_estampillas.text())
				
				
				#inicial y final corresponden a la gestion:
				#falta validar cantidad con un if: si ESTAMPILLA_FINAL-ESTAMPILLA_INICIAL > CANTIDAD
				inicial =ESTAMPILLA_INICIAL
				final =(ESTAMPILLA_INICIAL+cantidad)-1
				if self.txt_cantidad_estampillas.text():
					if self.rb_estampillas.isChecked():

						
						indice=4
						dav= int(self.txt_dav.text())
						categoria =str(self.txt_categoria_estampillas.currentText())
						variedad =str(self.txt_variedad_estampillas.text())
						envase=int(self.txt_envase_estampillas.text())
						#Validar DAV: la bandera valida si se encontró ese numero de dav (repetido)
						Ban=q.validaDav(dav)
						if Ban is False:
							q.carga_estampillas(rncyfs,dav,especie,categoria,campana,cantidad,variedad,inicial,final,envase,date.today())
						else:
							c.cartel("ERROR","DAV REPETIDO",3)
					
					else:

						indice=5
						q.carga_estampillas_anexo(rncyfs,especie,campana,cantidad,inicial,final,date.today())
				else:
					c.cartel("ERROR","INGRESE CANTIDAD",3)


				#DATOS PARA ACTUALIZAR RANGO GENERAL DE ESTAMPILLAS
				estampilla_siguiente=final+1
				ultima_estampilla=int(ESTAMPILLA_FINAL)
					

				q.actualizarangoenbd(estampilla_siguiente,ultima_estampilla,indice)
				self.txt_cantidad_estampillas.setText("")

				self.refresh_estampillas()
				self.ver_estampillas()

			else:

				c.cartel("ERROR","SELECCIONAR ASOCIADO",3)
		except Exception as e:
			c.cartel("ERROR",str(e),3)


	def traeRango(self):

		if self.rb_serieA_componer.isChecked():
				indice=1
		elif self.rb_serieB_componer.isChecked():
				indice=2
		elif self.rb_estampillas_componer.isChecked():
				indice = 4

		elif self.rb_anexo_componer.isChecked():
				indice=5
		
		rango_general=q.recuperarango(indice)
		self.txt_inicio_componer.setText(str(rango_general[0]))
		self.txt_final_componer.setText(str(rango_general[1]))
								   


	
	def componer_rango(self):
		if self.txt_inicio_componer.text() and self.txt_final_componer.text():
			inicio=int(self.txt_inicio_componer.text())
			fin = int (self.txt_final_componer.text())



			if self.rb_serieA_componer.isChecked():
				indice=1
			elif self.rb_serieB_componer.isChecked():
				indice=2
			elif self.rb_estampillas_componer.isChecked():
				indice = 4

			elif self.rb_anexo_componer.isChecked():
				indice=5
			
			
			r=c.cartel_opcion("ATENCION","DESEA CORREGIR LA NUMERACION",2)
			
			if r==16384:
				q.corregir_rango(indice,inicio,fin)
				c.cartel("ATENCION","RANGO MODIFICADO",1)
				
				self.txt_inicio_componer.setText("")
				self.txt_final_componer.setText("")
		else:
			c.cartel("ERROR","INGRESE RANGO INICIAL Y FINAL",3)

	def recupera_rango_pedido(self):
		num_pedido = int(self.txt_num_pedido_corregir.text())
		rango=q.recuperaPedido(num_pedido)
		self.txt_inicio_componer_pedido.setText(str(rango[0]))
		self.txt_final_componer_pedido.setText(str(rango[1]))
		self.txt_inicialRe_componer_pedido.setText(str(rango[2]))
		self.txt_finalRE_componer_pedido.setText(str(rango[3]))




	def componer_pedido(self):
		inicio =int(self.txt_inicio_componer_pedido.text())
		fin= int(self.txt_final_componer_pedido.text())
		inicioR = int(self.txt_inicialRe_componer_pedido.text())
		finR = int(self.txt_finalRE_componer_pedido.text())
		num_pedido = int(self.txt_num_pedido_corregir.text())

		r=c.cartel_opcion("ATENCION","DESEA CORREGIR LA NUMERACION",2)
		if r==16384:
			q.componer_rango_pedido(inicio,fin,inicioR,finR,num_pedido)
			c.cartel("ATENCION","RANGO MODIFICADO",1)

	
			
	def imprimirticket(self):
			os.startfile("ticket.txt", "print")
			
			
	def limpiar(self):
		self.txt_cantidad.setText("")
		self.txt_rncyfs.setText("")
		self.txt_kg.setText("")
		
		self.txt_subcantidad.setText("")
		self.txt_subvariedad.setText("")
		
		
	def ver_estampillas(self):

		if self.cbx_filtra_asociado.isChecked():
			asociado=str(self.combo_asociados_estampillas.currentText())
		else:
			asociado='%'
		
				


		if self.rb_estampillas.isChecked():

			if self.cbx_filtra_porfecha_estampillas.isChecked():
				desde=self.fecha_desde_estampillas.text()
				hasta=self.fecha_hasta_estampillas.text()
				listarecuperada=q.recuperaEstampillasFecha(asociado,desde,hasta)
			else:
				listarecuperada=q.recuperaEstampillas(asociado)
				


			
			totalfilas=len(listarecuperada)
			self.tb_estampillas.setRowCount(totalfilas)
				
				
			fila =0
			cantidad=0
			for i in listarecuperada:
				self.tb_estampillas.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_estampillas.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_estampillas.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_estampillas.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_estampillas.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_estampillas.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
				self.tb_estampillas.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
				self.tb_estampillas.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
				self.tb_estampillas.setItem(fila,8,QtWidgets.QTableWidgetItem(str(i[8])))
				self.tb_estampillas.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[9])))
				self.tb_estampillas.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[10])))
				self.tb_estampillas.setItem(fila,11,QtWidgets.QTableWidgetItem(str(i[11])))
				
					
				fila=fila+1
				cantidad=cantidad+int(i[4])
		else:

			listarecuperada=q.recuperaAnexos(asociado)
			totalfilas=len(listarecuperada)
			self.tb_estampillas.setRowCount(totalfilas)
			
				
			fila =0
			cantidad=0
			for i in listarecuperada:
				self.tb_estampillas.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_estampillas.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_estampillas.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_estampillas.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_estampillas.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_estampillas.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
				self.tb_estampillas.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
				self.tb_estampillas.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
				self.tb_estampillas.setItem(fila,8,QtWidgets.QTableWidgetItem(str(i[8])))
				self.tb_estampillas.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[9])))
				self.tb_estampillas.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[10])))
				self.tb_estampillas.setItem(fila,11,QtWidgets.QTableWidgetItem(str(i[11])))
				
					
				fila=fila+1
				cantidad=cantidad+int(i[4])
		self.signal_total_estampillas.setText(str(cantidad))


	
	def verpedidos(self):
		
		campo= str(self.txt_verpedido.text())
		
		
		tablapedidos=q.verpedido(campo)
		totalfilas=len(tablapedidos)
		self.tb_verpedidos.setRowCount(totalfilas)
		if totalfilas >0:
		
			fila=0
			
			
			for i in tablapedidos:
				
							
				self.tb_verpedidos.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_verpedidos.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_verpedidos.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_verpedidos.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))  
				
				
				fila=fila+1
		else:
			c.cartel("ERROR","EL ASOCIADO NO POSEE PEDIDOS VIGENTES",3)


	def ver_historialPedido(self,num_pedido):
		

		tablasubpedidos=q.recuperaSubpedidos(num_pedido)
		totalfilas=len(tablasubpedidos)
		self.tb_historial_subpedidos.setRowCount(totalfilas)
		if totalfilas >0:
		
			fila=0
			
			
			for i in tablasubpedidos:
				
							
				self.tb_historial_subpedidos.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_historial_subpedidos.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_historial_subpedidos.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_historial_subpedidos.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_historial_subpedidos.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_historial_subpedidos.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))  
				self.tb_historial_subpedidos.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
				self.tb_historial_subpedidos.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
				
				
				fila=fila+1

		self.signal_historial_subpedidos.setText(str(num_pedido))
		#self.tb_historial_subpedidos.verticalScrollBar().setSliderPosition(self.tb_historial_subpedidos.verticalScrollBar().maximum())

	
	def eliminar_subpedido(self):
		num_pedido =self.txt_numpedido.text()

		fila = self.tb_historial_subpedidos.currentRow()
		inicio =int(self.tb_historial_subpedidos.item(fila, 1).text())

		r=c.cartel_opcion("ATENCION","DESEA ELIMINAR EL SUBPEDIDO SELECCIONADO",2)
			
		if r==16384:
			q.eliminarLineaSubpedido(inicio,num_pedido)
			c.cartel("ATENCION","SUBPEDIDO ELIMINADO",1)

			q.corregirPedido(inicio,num_pedido)
		
		self.ver_historialPedido(num_pedido)
		self.verpedidos()
		self.refresh_pedidos()



		

			
	

	def completanumpedido(self):
		
		fila = self.tb_verpedidos.currentRow()
		pedido=self.tb_verpedidos.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		self.txt_numpedido.setText(pedido)
		self.validarpedido()
		
		self.ver_historialPedido(int(pedido))


		
		
	def refresh_pedidos(self):
		self.validarpedido()
		self.verpedidos()	
				
	def validarpedido(self):
		numeropedido=int(self.txt_numpedido.text())
		
		tablapedidos=q.getpedido(numeropedido)
		
		if len(tablapedidos) != 0:
			disponible=int(tablapedidos[7]-tablapedidos[6]+1)
		
			
			self.signal_disponibleinicio_sp.setText(str(tablapedidos[6]))
			self.signal_disponiblefin_sp.setText(str(tablapedidos[7]))
			
			self.signal_stock_sp.setText(str(disponible))
		
		
			
	def clipinicio(self):
		inicio=int(self.signal_disponibleinicio_sp.text())
		clip.copy(inicio)
		
		
	def nuevosubpedido(self):
		
		
		if self.txt_subvariedad.text() and self.txt_subcantidad.text() and self.txt_kg.text() and self.cbx_especie_sp.currentText() and self.cbx_categoria_sp.currentText(): 
		
			numeropedido=int(self.txt_numpedido.text())
			
			tablapedidos=q.getpedido(numeropedido)
			cantidad =int(self.txt_subcantidad.text())
			
			if len(tablapedidos) != 0:
				disp=int(tablapedidos[7]-tablapedidos[6]+1)
				
				if disp >= cantidad:
					
					registro=str(tablapedidos[1])
						
					
						
					spini=int(tablapedidos[6]) #valor inicioremanente de pedido
					spfin=spini+cantidad-1
					
					numpedido=int(self.txt_numpedido.text())
					variedad=str(self.txt_subvariedad.text())
					especie=str(self.cbx_especie_sp.currentText())
					KG=int(self.txt_kg.text())
					categoria=str(self.cbx_categoria_sp.currentText())
					camp=self.txt_subcamp.text()
					fechasubpedido=str(date.today())
					#fechasubpedido=formatearfecha(fecha)
										
					
					if variedad=="":
						variedad="n/d"
					if especie=="n/d":
						especie="n/d"
					
					if categoria=="":
						categoria="n/d"
					if KG=="":
						dav=0
						
							
					
				
					'''#EXPORTA A ARCHIVO EXCELL
					book = Workbook()
					sheet = book.active
					
					
					sheet['I13']=str(spini)+" - "+str(spini+cantidad-1)
				
					
				
				
					book.save('nota_davs.xls',"a")'''
					
					
					c.cartel("AVISO","SUBPEDIDO CREADO",1)
					
					clip.copy(str(spini)+"-"+str(spfin))
					
					q.cargasubepedido(numpedido,spini,spfin,cantidad,variedad,especie,int(camp),int(KG),categoria,registro,fechasubpedido)
					
					inicioremanente=tablapedidos[6]+cantidad
					disponible=tablapedidos[7]-inicioremanente+1
					if disponible <=0:
						estado="FINALIZADO"
					else:
						estado="VIGENTE"
						
					q.actualizaremanente(numpedido,inicioremanente)				 #se actualiza el stock remanente del pedido en tabla pedidos
					q.actualizaestado(numpedido,estado)
					self.refresh_pedidos()
					self.ver_historialPedido(numpedido)
				else:
				
					c.cartel("ERROR DE STOCK","NO HAY STOCK SUFICIENTE PARA ESTA SOLICITUD",3)
		else:
		
			c.cartel("ERROR","CAMPOS VACIOS",3)
				
	def deshacerSubpedido(self):
		'''
		num_pedido=int(self.txt_num_deshacer.text())
		fecha_sub= str (self.txt_fecha_deshacer.text())
		lista_recuperada=q.traeSubpedido(num_pedido,fecha_sub)
		primer_rango=q.traeRangoInicial(num_pedido,fecha_sub)
		cantidad=0
		
		for i in lista_recuperada:
			
			cantidad=int(i[1])+cantidad
		
		
		q.deshacer(num_pedido,fecha_sub,int(cantidad),int(primer_rango[0]))
		
		
			
			
				
			
				
				
				
			
			
			
			
			
		#recuperar cantidad de rotulos y rango inicial
		#restablecer agregar la cantidad al stock y volver al rotulo de inicio
					
		'''	
			
	def altasocio(self):
		
		registro =str(self.txt_altaasociado_reg.text())
		nombre =str(self.txt_altaasociado_nombre.text())
		resultado=q.validarasociado(registro)
		valor=int("".join(map(str,resultado)))
		print (valor)
		if valor ==1:
			
			c.cartel("ATENCION","YA EXISTE ESE ASOCIADO",3)
		else:
					
			q.altaasociado(registro,nombre)
			
			self.txt_altaasociado_reg.setText("")
			self.txt_altaasociado_nombre.setText("")
			c.cartel("INFORMACION","ASOCIADO REGISTRADO",1)
		
		
		
	def nuevagestion(self):
		
		
		if self.txt_cantidad_gestiones.text():
			
			if self.txt_rncyfs_gestiones.text():
				registro=str(self.txt_rncyfs_gestiones.text())
				
			else:
				registro =self.traeregistrogestiones()
			
			tipo =str(self.combo_gestiones.currentText())
			cantidad =int(self.txt_cantidad_gestiones.text())
			
			fechagestion=str(date.today())
			#fechagestion=formatearfecha(fecha)
			
			
			if tipo == "Envio Inspecciones" or tipo== "Alta DT" or tipo =="Otra Gestion":
				estado="PENDIENTE"
				
			elif tipo =="Subpedido":
				
				
				estado="EN IMPRESION"
				num_gestion=q.traeIndiceGestion()
				q.altarotulo(registro,"Soja","IQR",cantidad,"PENDIENTE","Primera",fechagestion,int(num_gestion[0]))
				
			else:
				estado="PENDIENTE"
				
			
			count=q.validarasociado(registro)
			valor=int("".join(map(str,count)))
			if valor ==0:
				
				c.cartel("ERROR","NO EXISTE ESE ASOCIADO",3)
			else:	
			
				
				
				q.altagestion(registro,tipo,estado,fechagestion,cantidad)
				
				self.combo_gestiones.setCurrentIndex(0)
				self.combo_asociados_gestiones.setCurrentIndex(0)
				self.txt_cantidad_gestiones.setText("")
				self.txt_rncyfs_gestiones.setText("")
				self.listargestiones()
				
			
				c.cartel("AVISO","GESTION AGREGADA",1)
			
		else:
			
				
				
				
			
			c.cartel("ERROR","INGRESE CANTIDAD",3)
		
		
		
		
	def listargestiones(self):
		
		nombre= "%"
		estado = str(self.combo_filtro_estados.currentText())
		if estado == "-":
			estado="%"
			
		if self.rb_activas.isChecked():
			listarecuperada=q.traergestionesActivas(estado,nombre)
		else:
					
			listarecuperada=q.traergestiones(estado,nombre)
		
		
		
		
		
		totalfilas=len(listarecuperada)
		self.tb_gestiones.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_gestiones.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_gestiones.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_gestiones.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_gestiones.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_gestiones.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
			self.tb_gestiones.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
			self.tb_gestiones.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
									
			fila=fila+1
		self.signal_gestiones.setText(str(totalfilas))
		
	def filtrargestiones(self):
		if str(self.combo_estados_gestiones_filtro.currentText()) =="-":
			
			estado="%"
		else:
			estado=str(self.combo_estados_gestiones_filtro.currentText())
			
		asociado=str("%"+self.txt_asociados_gestiones_filtro.text()+"%")
	
		
		
		listarecuperada=q.traergestiones(estado,asociado)
		totalfilas=len(listarecuperada)
		self.tb_gestiones.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_gestiones.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_gestiones.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_gestiones.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_gestiones.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_gestiones.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
			self.tb_gestiones.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
			
									
			fila=fila+1
		self.signal_gestiones.setText(str(totalfilas))
		
			
	def gestionselected(self):
		fila = self.tb_gestiones.currentRow()
		indice=self.tb_gestiones.item(fila, 6).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		registro=self.tb_gestiones.item(fila, 0).text()
		fecha=self.tb_gestiones.item(fila, 5).text()
		estado=self.tb_gestiones.item(fila, 1).text()
		
		
		obs=q.traenotas(int(indice))
		
		self.signal_gestion_asociado.setText(str(registro))
		self.signal_gestion_fecha.setText(str(fecha))
		self.signal_gestion_estado.setText(str(estado))
		self.signal_gestion_indice.setText(str(indice))
		self.signal_gestion_observaciones.setText(str("".join(obs)))
		
	def gestionupdate(self):
		indice=int(self.signal_gestion_indice.text())
		estado=str(self.combo_estados_gestiones.currentText())
		
		q.updategestion(indice,estado)
		
		self.combo_estados_gestiones.setCurrentIndex(0)
		self.listargestiones()
		
	def notaupdate(self):
		indice=int(self.signal_gestion_indice.text())
		obs=str(self.signal_gestion_observaciones.toPlainText())
		
		q.insertarnota(indice,obs)
		self.listargestiones()
		
		c.cartel("NOTA","NOTA AGREGADA",1)
	
	def eliminargestion(self):
		
		indice=int(self.signal_gestion_indice.text())
		
		r=c.cartel_opcion("ATENCION","DESEA ELIMINAR LA GESTION",2)
		
		if r==16384:
			q.borrargestion(indice)
			c.cartel("ATENCION","REGISTRO ELIMINADO",1)
			self.gestionupdate()
						
		
			
			
	def eliminarimpresion(self):
		
		indice=int(self.txt_indice_rotulos.text())
		r=c.cartel_opcion("ATENCION","DESEA ELIMINAR LA IMPRESION",2)
		
		
		
		if r==16384:
			q.borrarimpresion(indice)
			
			c.cartel("ATENCION","REGISTRO ELIMINADO",1)
			self.listarimpresiones()
						
	def eliminar_envio(self):
		
		indice=int(self.signal_id_envio.text())
		r=c.cartel_opcion("ATENCION","DESEA ELIMINAR EL ENVIO SELECCIONADO",2)
		
		
		
		if r==16384:
			q.eliminarEnvio(indice)
			
			c.cartel("ATENCION","ENVIO ELIMINADO",1)
			self.traerenvios()
			self.signal_envio_fecha.setText("")
			self.signal_envio_estado.setText("")
			self.signal_envio_guia.setText("")
			self.signal_id_envio.setText("")
			self.signal_envio_asociado.setText("")
			self.signal_envio_fct.setText("")
			
			
	def nuevoEnvio(self):

		fecha_envio=date.today()
		registro=str(q.getrncyfs(str(self.signal_asociado_envios.text()))[0])
		cantidad=int(self.txt_cantidad_envios.text())
		especie=str(self.txt_especie_envios.text())
		bultos =str(self.txt_bultos_envios.text())
		detalle=str(self.cb_detalle_envios.currentText())
		obs=str(self.txt_obs_envios.text())
		fecha_emision= str(self.signal_emision.text())
		estado='PREPARADO'
		tipo=str(self.cb_servicio_envios.currentText())
		
		
		if self.cbx_incluye.isChecked():
			rotulos='SI'
		else:
			rotulos='NO'
		
		if tipo and bultos:
			q.insertarEnvio(fecha_envio,registro,cantidad,tipo,rotulos,fecha_emision,bultos,estado,especie,detalle,obs)
			self.txt_cantidad_envios.setText("")
			self.txt_especie_envios.setText("")
			self.txt_bultos_envios.setText("")
			self.cb_detalle_envios.setCurrentIndex(0)
			self.txt_obs_envios.setText("")
			self.signal_emision.setText("")
			self.cb_servicio_envios.setCurrentIndex(0)
		
			c.cartel("AVISO","ENVIO CREADO",1)
			self.traerenvios()
		else:
			c.cartel("ERROR","FALTAN CAMPOS",3)


	def agregar_guia(self):
		guia= str(self.txt_guia_envio.text())
		indice=int(self.signal_id_envio.text())
		
		q.agregarGuia(guia,indice)
		self.signal_envio_guia.setText(guia)
		self.txt_guia_envio.setText("")

	def agregar_factura(self):
		factura= str(self.txt_fct_envios.text())
		indice=int(self.signal_id_envio.text())
		
		q.agregarFct(factura,indice)
		self.signal_envio_fct.setText(factura)
		self.txt_fct_envios.setText("")

		
		
	def calcular_costo_sugerido(self):
		costo=float(self.txt_costo_envio.text()	)
		precio_sugerido=round(costo*1.3)
		
		digito=int(len(str(precio_sugerido)))
		multiplica=0
		
		if digito == 4:
			multiplica = (digito-1) *100

		elif digito == 3:
			multiplica = (digito-1) *100
		elif digito == 2:
			multiplica = (digito-1) *10
		elif digito ==1:
			multiplica= (digito-1) *1

		
			



		self.txt_precio_sugerido.setText(str(precio_sugerido))
		
		
	def definir_envio_facturado(self):
		indice=int(self.signal_id_envio.text())
		q.definirEnvioFacturado(indice)
		c.cartel("ESTADO ENVIO","ENVIO FACTURADO",1)
		self.traerenvios()
		
			
	def listar(self):
		
		
		if self.cbx_porrotulo.isChecked():
				try:
					rotulo=int(self.txt_porrotulo.text())
					tablarecuperada = q.busquedaxrotulo(rotulo)
					
					i=0
					for i in tablarecuperada:
						self.signal_pedido_bxr.setText(str(i[0]))
						numpedido=int(i[0])
						break
					razon_social=q.traerazonsocial(numpedido)
					
					self.signal_asociado_bxr.setText("".join(razon_social))
				except Exception as e:
					c.cartel("PEDIDO NO ENCONTRADO","No se encontro el pedido correspondiente a ese rotulo",3)
			
			
		else:
			
		
				
			if self.rb_vigentes.isChecked():
				radiobutton="VIGENTE"
			elif self.rb_finalizados.isChecked():
				radiobutton="FINALIZADO"
			elif self.rb_todos.isChecked():
				radiobutton="%"			
			elif self.rb_sinusar.isChecked():
				radiobutton="SIN USAR"
				
			desde=str(self.fecha_desde_listar.text())
			hasta=str(self.fecha_hasta_listar.text())
		
		
		
			
			if self.rb_porrncyfs.isChecked():
				campolistar= str(self.txt_listar.text())
				
				if self.cbx_porfecha.isChecked():
					
					tablarecuperada=q.listaxregistrofecha(campolistar,radiobutton,desde,hasta)
				else:
					tablarecuperada=q.listaxregistro(campolistar,radiobutton)
				
				
				
			
				
			elif self.rb_pornumpedido.isChecked():
				campolistar= int(self.txt_listar.text())
				
				if self.cbx_porfecha.isChecked():
					tablarecuperada=q.listaxpedidofecha(campolistar,radiobutton,desde,hasta)
				else:
					tablarecuperada=q.listaxpedido(campolistar,radiobutton)
				
			
		
		totalfilas=len(tablarecuperada)
		self.tb_listar.setRowCount(totalfilas)		
			
		fila=0
		
		acum=0
		for i in tablarecuperada:
			
						
			self.tb_listar.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_listar.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_listar.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2]))) #RAZON SOCIAL
			self.tb_listar.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_listar.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
			self.tb_listar.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
			self.tb_listar.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
			self.tb_listar.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
			self.tb_listar.setItem(fila,8,QtWidgets.QTableWidgetItem(str(i[8])))
			self.tb_listar.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[9])))
			self.tb_listar.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[10])))
			self.tb_listar.setItem(fila,11,QtWidgets.QTableWidgetItem(str(i[11])))
			self.tb_listar.setItem(fila,12,QtWidgets.QTableWidgetItem(str(i[12])))
			
		
			acum=acum+int(i[4])
				
			fila+=1
			
		self.signal_total_listar.setText(str(acum))
			
	def listartoexcell(self):
		
		book = Workbook()
		sheet = book.active
		
		
		
		
		
		if self.rb_pornumpedido.isChecked():
			campolistar= int(self.txt_listar.text())
			tablarecuperada=q.listaraexcell(campolistar)
			#TRAER RAZON SOCIAL CON NUM DE PEDIDO O CON NUM DE REGISTRO
		else:
			campolistar= str(self.txt_listar.text())
			tablarecuperada=q.listaraexcellall(campolistar)
		
		
		sheet['A1']="ASOCIADO"
		sheet['B1']=""
		sheet['A2']="PEDIDO"
		sheet['B2']="CANTIDAD"
		sheet['C2']="RANGO UTILIZADO"
		sheet['D2']="VARIEDAD"
		sheet['E2']="ESPECIE"
		sheet['F2']="CATEGORIA"
		sheet['G2']="ENVASES/KG"
		
		sheet['H2']="CONTROL"
		
		for i in tablarecuperada:
			sheet.append(i)
			
					
		path = "detalle_{}.xlsx".format(date.today())
							
			
		book.save(path)
		c.cartel("AVISO","DETALLE EXPORTADO",1)
			
		
	def rendir(self):
		if self.rb_todo_rendicion.isChecked():
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			especie=str(self.cbx_especie_rendicion.currentText())
			cultivar=str(self.cbx_cultivar_rendicion.currentText())
			categoria=str(self.cbx_categoria_rendicion.currentText())
			camp= str(self.cbx_camp_rendicion.currentText())
			envase=str(self.cbx_envase_rendicion.currentText())
		
		
		
		
			if registro =="":
				registro='%'
			
			if especie =="":
				especie='%'
				
			if cultivar =="":
				cultivar='%'
		
			if categoria =="":
				categoria='%'

			if envase=='40 KG.':
				envase=40
				listarecuperada =q.listarrendicion40(desde,hasta,registro,especie,cultivar,categoria,camp,envase)
			elif envase =="Big Bag":
				
				listarecuperada =q.listarrendicionBB(desde,hasta,registro,especie,cultivar,categoria,camp)
			else:
				listarecuperada =q.listarrendicion(desde,hasta,registro,especie,cultivar,categoria,camp)
				

			
			
			
			
				
			
			
			totalfilas=len(listarecuperada)
			self.tb_rendicion.setRowCount(totalfilas)
		
				
			fila =0
			acum=0
			for i in listarecuperada:
				self.tb_rendicion.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_rendicion.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_rendicion.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_rendicion.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_rendicion.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_rendicion.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
				self.tb_rendicion.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
				self.tb_rendicion.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
				self.tb_rendicion.setItem(fila,8,QtWidgets.QTableWidgetItem(str(i[8])))
				self.tb_rendicion.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[9])))
				self.tb_rendicion.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[10])))
				fila=fila+1
				acum=acum+int(i[1])
				
			self.signal_total_rendicion.setText(str(acum))
			
		else:
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			
			if registro =="":
				registro="%"
			
			
			
			listarecuperada =q.listarrendicionsolopedidos(desde,hasta,registro)
			totalfilas=len(listarecuperada)
			self.tb_rendicion.setRowCount(totalfilas)
			
			
			fila =0
			acum=0
			for i in listarecuperada:
				self.tb_rendicion.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_rendicion.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_rendicion.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_rendicion.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_rendicion.setItem(fila,4,QtWidgets.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,5,QtWidgets.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,6,QtWidgets.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,7,QtWidgets.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,8,QtWidgets.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_rendicion.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[5])))
				
				fila=fila+1
				acum=acum+int(i[1])
			
			
			
			self.signal_total_rendicion.setText(str(acum))
				
			
		
	def exportarrendicion(self):
		
		if self.rb_todo_rendicion.isChecked():
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			especie=str(self.cbx_especie_rendicion.currentText())
			cultivar=str(self.cbx_cultivar_rendicion.currentText())
			categoria=str(self.cbx_categoria_rendicion.currentText())
			camp= str(self.cbx_camp_rendicion.currentText())
		
		
		
		
			if registro =="":
				registro="%"
			
			if especie =="":
				especie="%"
				
			if cultivar =="":
				cultivar="%"
		
			if categoria =="":
				categoria="%"
			
			if camp =="":
				camp="%"
			
				
			
			tablarecuperada =q.listarrendicion(desde,hasta,registro,especie,cultivar,categoria,camp)
			book = Workbook()
			sheet = book.active
		
			sheet['A1']="RENDICION MENSUAL"
			sheet['A2']="PERIODO"
			sheet['B2']=desde +" a "+hasta	
			sheet['D2']="CANTIDAD ASIGNADA"	
			sheet['E2']="=SUM(B3:B1000)"
				
			sheet['A3']="RANGO"
			sheet['B3']="CANTIDAD"
			sheet['C3']="RNCYFS"
			sheet['D3']="RAZON SOCIAL"
			sheet['E3']="KG"
			sheet['F3']="ESPECIE"
			sheet['G3']="CULTIVAR"
			sheet['H3']="CATEGORIA"
			sheet['I3']="CAMPANA"
			sheet['J3']="FECHA"
			sheet['K3']="PEDIDO N"
		
			for i in tablarecuperada:
				sheet.append(i)
			
					
						
			
			book.save('rendicion_subpedidos.xls')
			
			c.cartel("RENDICION","RENDICION GENERADA",1)
			
			
		else:
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			
			if registro =="":
				registro="%"
			
			
			
			tablarecuperada =q.listarrendicionsolopedidos(desde,hasta,registro)
			book = Workbook()
			sheet = book.active
		
			sheet['A1']="RENDICION MENSUAL"
			sheet['A2']="PERIODO"
			sheet['B2']=desde +" a "+hasta
			sheet['D2']="CANTIDAD ENTREGADA"	
			sheet['E2']="=SUM(B3:B500)"
			
				
			sheet['A3']="RANGO"
			sheet['B3']="CANTIDAD"
			sheet['C3']="RNCYFS"
			sheet['D3']="RAZON SOCIAL"
			sheet['E3']="FECHA"
			sheet['F3']="PEDIDO N"
		
			for i in tablarecuperada:
				sheet.append(i)
			
					
						
			
			book.save('rendicion_pedidos.xls')
			c.cartel("DETALLE","RENDICION GENERADA",1)
		
					
			
		
	def setearlockers(self):

		r=c.cartel_opcion("ATENCION","DESEA INICIALIZAR LOS LOCKERS",2)
			
		if r==16384:
			cantidad=self.txt_definir_locker.text()
		
			
			for i in range(1,int(cantidad)):
				estado ="Disponible"
				q.definircantidadlockers(i,estado)
							
				

		
	
			
	def lockerdisponibles(self):
		
		lockers=q.recuperalockers()
		
		#la consulta devuelve una tupla, por lo tanto hay que convertirla a str para llenar el combobox, se usa el metodo "".join()
		
		k=0
		
		for i in lockers:
			self.cbx_num_locker.addItem("".join(map(str,lockers[k]))) #uso funcion map para pasar de tupla entero a string
			k=k+1
		
	def almacenar(self):
		
		if self.txt_pedido_deposito.text():
			
			locker = str(self.cbx_num_locker.currentText())
			pedido =int(self.txt_pedido_deposito.text())
			
			fechaingreso= str(date.today())
			#fechaingreso=formatearfecha(fecha)
			
			q.modificalocker(locker,pedido,fechaingreso)
			
			self.cbx_num_locker.clear()
			self.lockerdisponibles()
			self.listarlockers()
			c.cartel("AVISO","PEDIDO ALMACENADO",1)
			
		else:
			
			c.cartel("ERROR","INGRESE PEDIDO A ALMACENAR",3)
		
		
	def lockerselected(self):
		fila = self.tb_lockers.currentRow()
		locker=self.tb_lockers.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		pedido=self.tb_lockers.item(fila, 1).text()
		nombre=self.tb_lockers.item(fila, 5).text()
		
		
		self.signal_pedido_locker.setText(str(pedido))
		self.signal_locker_locker.setText(str(locker))
		self.signal_razon_locker.setText(str(nombre))
			
		
	def despachar(self):
		locker = int(self.signal_locker_locker.text())
		
		
		
		q.liberalocker(locker)
		c.cartel("AVISO","LOCKER LIBERADO",1)
		
		self.cbx_num_locker.clear()
		self.lockerdisponibles()
		self.listarlockers()
		
		
			
	def listarlockers(self):
		
		listarecuperada=q.verlockers()
		totalfilas=len(listarecuperada)
		self.tb_lockers.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_lockers.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_lockers.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_lockers.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_lockers.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_lockers.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
			self.tb_lockers.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
			
				
			fila=fila+1
			
	def filtrarlockers(self):
		
		locker=str("%"+self.txt_buscar_locker.text()+"%")
		listarecuperada=q.verlockers_filtrado(locker)
		totalfilas=len(listarecuperada)
		self.tb_lockers.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_lockers.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_lockers.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_lockers.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_lockers.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_lockers.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
			self.tb_lockers.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
			
				
			fila=fila+1
		
		
	def nuevaimpresion(self):
		
		
		registro=str(self.txt_rncyfs_rotulos.text())
		count=q.validarasociado(registro)
		valor=int("".join(map(str,count)))
		
		if valor ==0:
			
			c.cartel("ERROR","NO EXISTE ESE ASOCIADO",3)
		else:
					
			if self.txt_cantidad_rotulos.text():
				
				cantidad=int(self.txt_cantidad_rotulos.text())
				especie=str(self.cbx_especie_rotulos.currentText())
				categoria=str(self.cbx_categoria_rotulos.currentText())
				tipo=str(self.cbx_tipo_rotulos.currentText())
				
				
				fechaimpresion=str(date.today())
				#fechaimpresion=formatearfecha(fecha)
				estado="PENDIENTE"
				
				indice=0
				if tipo == "Tyveck 40KG.":
					if categoria =="Primera":
						indice=1
						
					elif categoria =="Original":
						indice=4
					elif categoria == "Segunda":
						indice=3
					elif categoria =="Identificada":
						indice =5
						
				elif tipo == "Tyveck BB":
					if categoria =="Primera":
						indice=2
					elif categoria =="Identificada":
						indice=6
				
			
					
				if indice !=0:
					
					stock=q.recuperastock(indice)
					stock_actualizado = int(stock[0])-cantidad
					if stock_actualizado <0:
						
						c.cartel("ERROR","NO HAY STOCK SUFICIENTE DE ESE TIPO DE ROTULO",3)
					else:
						q.actualizar_stock(stock_actualizado,indice)
						q.altarotulo(registro,especie,tipo,cantidad,estado,categoria,fechaimpresion,0)
						self.traerstock()
						c.cartel("AVISO","PEDIDO DE IMPRESION INGRESADO",1)						
					
					
					
				
				else:
					q.altarotulo(registro,especie,tipo,cantidad,estado,categoria,fechaimpresion,0)
					c.cartel("AVISO","PEDIDO DE IMPRESION INGRESADO",1)
				
				
					
					
				self.txt_rncyfs_rotulos.setText("")
				self.txt_cantidad_rotulos.setText("")
				self.cbx_especie_rotulos.setCurrentIndex(0)
				self.cbx_categoria_rotulos.setCurrentIndex(0)
				self.cbx_tipo_rotulos.setCurrentIndex(0)
				self.listarimpresiones()
				
			else:
				
				c.cartel("ERROR","INGRESE CANTIDAD",3)
		
	def cargaStock(self):
		
		if self.txt_cantidad_rotulos_stock.text():
			
			tipo=str(self.cbx_tipo_rotulos_ingresar.currentText())
			cantidad =int(self.txt_cantidad_rotulos_stock.text())
			categoria =str(self.cbx_categoria_rotulos_stock.currentText())
			
			if tipo == "Tyveck 40KG.":
						if categoria =="Primera":
							indice=1
							
						elif categoria =="Original":
							indice=4
						elif categoria == "Segunda":
							indice=3
						elif categoria =="Identificada":
							indice =5
							
			elif tipo == "Tyveck BB":
						if categoria =="Primera":
							indice=2
						elif categoria =="Identificada":
							indice=6
			
			stock_recuperado=q.recuperastock(indice)	
			stock =int(stock_recuperado[0])
			stock_final=stock+cantidad
			q.actualizar_stock(stock_final,indice)
			
			c.cartel("AVISO","STOCK ACTUALIZADO",1)
			
	
			self.traerstock()
			
			self.txt_cantidad_rotulos_stock.setText("")
		else:
		
			c.cartel("ERROR","INGRESE CANTIDAD",3)
			
		
	def limpiar_formulario_busqueda_impresiones(self):
		self.fecha_desde_rotulos.setDate(date.today())
		self.fecha_hasta_rotulos.setDate(date.today())
		self.cb_razonsocial.setCurrentIndex(0)
		self.cb_tipo.setCurrentIndex(0)
		self.cb_especie.setCurrentIndex(0)
		self.cb_categoria.setCurrentIndex(0)
		self.rb_rotulos_pendientes.setChecked(True)


		
	def listarimpresiones(self):
		
		if self.rb_rotulos_pendientes.isChecked():
			estado="PENDIENTE"
		elif self.rb_rotulos_facturados.isChecked():
			estado="FACTURADO"
		elif self.rb_rotulos_todos.isChecked():
			estado="%"
		elif self.rb_rotulos_completos.isChecked():
			estado="COMPLETO"
		elif self.rb_rotulos_pendientesdav.isChecked():
			estado="PENDIENTE DAV"
			
			
		if self.cb_tipo.currentText() == "-":
			tipo="%"
		else:
				
			tipo=str(self.cb_tipo.currentText())
				
		if self.cb_especie.currentText()=="-":
			especie="%"
		else:
			especie=str(self.cb_especie.currentText())
			
		if self.cb_razonsocial.currentText()=="-":
			razon="%"
		else:
			razon=str(self.cb_razonsocial.currentText())
		
		
		if self.cb_alldate.isChecked():
			
			listarecuperada=q.traerotulos(estado,tipo,especie,razon)
		else:
			inicio=str(self.fecha_desde_rotulos.text())
			fin=str(self.fecha_hasta_rotulos.text())
			listarecuperada=q.traerotulosFecha(estado,tipo,especie,inicio,fin)
			
		
		
		totalfilas=len(listarecuperada)
		self.tb_rotulos.setRowCount(totalfilas)
			
			
		fila =0
		acum=0
		for i in listarecuperada:
			self.tb_rotulos.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[0]))) #TIPO DE ROTULO
			self.tb_rotulos.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1]))) #FECHA
			self.tb_rotulos.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2]))) #ESTADO
			if i[2] =="FACTURADO":
						
					self.tb_rotulos.setItem(fila, 2, QtWidgets.QTableWidgetItem("FACTURADO"))
					self.tb_rotulos.item(fila,2).setForeground(QtGui.QColor(205,221,193))
			elif i[2] =="PENDIENTE":
					self.tb_rotulos.setItem(fila, 2, QtWidgets.QTableWidgetItem("PENDIENTE"))
					self.tb_rotulos.item(fila, 2).setForeground(QtGui.QColor(254, 247, 105))
			elif i[2] =="COMPLETO":
					self.tb_rotulos.setItem(fila, 2, QtWidgets.QTableWidgetItem("COMPLETO"))
					self.tb_rotulos.item(fila,2).setForeground(QtGui.QColor(148,178,214))
					
			self.tb_rotulos.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3]))) #CANTIDAD
			self.tb_rotulos.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4]))) #RAZON SOCIAL
			self.tb_rotulos.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5]))) #ESPECIE
			self.tb_rotulos.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6]))) #CATEGORIA
			if i[6] =="Primera":
						
					self.tb_rotulos.setItem(fila, 6, QtWidgets.QTableWidgetItem("Primera"))
					self.tb_rotulos.item(fila,6).setBackground(QtGui.QColor(255,162,154))
			elif i[6] =="Segunda":
					self.tb_rotulos.setItem(fila, 6, QtWidgets.QTableWidgetItem("Segunda"))
					self.tb_rotulos.item(fila,6).setBackground(QtGui.QColor(167,217,251))
			elif i[6] =="Identificada":
					self.tb_rotulos.setItem(fila, 6, QtWidgets.QTableWidgetItem("Identificada"))
					self.tb_rotulos.item(fila,6).setBackground(QtGui.QColor(245,242,170))
			elif i[6] =="Original":
					self.tb_rotulos.setItem(fila, 6, QtWidgets.QTableWidgetItem("Original"))
					self.tb_rotulos.item(fila,6).setBackground(QtGui.QColor(255,255,255))
			
			
			
			
			self.tb_rotulos.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[7]))) #INDICE
			
			acum=acum+int(i[3])
				
			fila=fila+1
			
			self.signal_total_rotulos.setText(str(acum))

	def exportar_listado_rotulos(self):
		
		
		if self.rb_rotulos_pendientes.isChecked():
			estado="PENDIENTE"
		elif self.rb_rotulos_facturados.isChecked():
			estado="FACTURADO"
		elif self.rb_rotulos_todos.isChecked():
			estado="%"
		elif self.rb_rotulos_completos.isChecked():
			estado="COMPLETO"
		elif self.rb_rotulos_pendientesdav.isChecked():
			estado="PENDIENTE DAV"
			
			
		if self.cb_tipo.currentText() == "-":
			tipo="%"
		else:
				
			tipo=str(self.cb_tipo.currentText())
				
		if self.cb_especie.currentText()=="-":
			especie="%"
		else:
			especie=str(self.cb_especie.currentText())
			
		if self.cb_razonsocial.currentText()=="-":
			razon="%"
		else:
			razon=str(self.cb_razonsocial.currentText())
		
		
		if self.cb_alldate.isChecked():
			
			listarecuperada=q.traerotulos(estado,tipo,especie,razon)
		else:
			inicio=str(self.fecha_desde_rotulos.text())
			fin=str(self.fecha_hasta_rotulos.text())
			listarecuperada=q.traerotulosFecha(estado,tipo,especie,inicio,fin)

		book = Workbook()
		sheet = book.active
		
		sheet['A1']="ID"
		sheet['B1']="FECHA"
		sheet['C1']="ESTADO"
		sheet['D1']="CANTIDAD"
		sheet['E1']="RAZON SOCIAL"
		sheet['F1']="ESPECIE"
		sheet['G1']="CATEGORIA"
		sheet['H1']="TIPO"
		
		
		for i in listarecuperada:
			sheet.append(i)
		
				
		book.save('rotulos_solicitados.xlsx')

		c.cartel("INFORMACION","PLANILLA CREADA",1)

		
			
	def cambiarestadorotulo(self):
		indice=int(self.txt_indice_rotulos.text())
		estado=str(self.txt_estado_rotulos.currentText())
		
		q.definirestadorotulo(estado,indice)
		
		#tambien actualizo estado en Gestiones
		num_gestion=q.traerIDgestion(indice)
		
		if num_gestion[0] != None:
			if estado =="COMPLETO":
			
				q.updategestion(int(num_gestion[0]),"DAV GESTIONADO")
			elif estado =="FACTURADO":
				q.updategestion(int(num_gestion[0]),"FINALIZADO")
			else:
				q.updategestion(int(num_gestion[0]),estado)
				
		c.cartel("INFORMACION","ESTADO MODIFICADO",1)
				
		self.listarimpresiones()
		
			
		
	def editarcantidadrotulo(self):
		# * * * atencion: a este modulo le falta desarrollar validacion sobre si correponde actualizar stock si se modifican las cantidades * * *
		
		if self.signal_rotulos_cantidad.text():
			r=c.cartel_opcion("ATENCION","DESEA MODIFICAR LA CANTIDAD",2)
			
			if r==16384:
						
				indice=int(self.txt_indice_rotulos.text())
				cantidad=int(self.signal_rotulos_cantidad.text())
				
				q.modificarCantidadImpresion(indice,cantidad)
				self.listarimpresiones()
				c.cartel("IMPRESION","CANTIDAD MODIFICADA",1)
		else:
			c.cartel("ATENCION","NO HAY NADA QUE MODIFICAR",3)

	def modificar_tipo(self):
		
		
		indice=int(self.txt_indice_rotulos.text())
		tipo=str(self.cb_corregir_tipo.currentText())
				
		q.modificarTipo(indice,tipo)
		self.listarimpresiones()
		c.cartel("TIPO","TIPO DE ROTULOS MODIFICADO",1)
		
			
			
		
		
		
		
		
	def impresionselected(self):
		fila = self.tb_rotulos.currentRow()
		indice=self.tb_rotulos.item(fila, 7).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		razon=self.tb_rotulos.item(fila, 4).text()
		cantidad=self.tb_rotulos.item(fila, 3).text()
		
		
		
		
		
		self.signal_rotulos_razon.setText(str(razon))
		self.signal_rotulos_cantidad.setText(str(cantidad))
		self.txt_indice_rotulos.setText(str(indice))
		
			
		
	def agregarrango(self):
			inicio=int(self.txt_inicio_nuevorango.text())
			cantidad=int(self.txt_fin_nuevorango.text())
			if self.rb_seriea_add.isChecked():
				serie="A"
			elif self.rb_serieb_add.isChecked():
				serie="B"
			fin=inicio+cantidad-1
			
			q.nuevorango(inicio,fin,cantidad,serie)
			
			self.txt_inicio_nuevorango.setText("")
			self.txt_fin_nuevorango.setText("")
			
			
	def verrangos(self):
		
		if self.rb_rangos_disponibles.isChecked():
			estado="DISPONIBLE"
		elif self.rb_rangos_enuso.isChecked():
			estado="EN USO"
		elif self.rb_rangos_terminados.isChecked():
			estado="TERMINADOS"
		elif self.rb_rangos_todos.isChecked():
			estado="%"
		
		if self.rb_definir_seriea.isChecked():
			serie="A"
		else:
			serie="B"
		
		
		listarecuperada=q.traerangos(estado,serie)
		totalfilas=len(listarecuperada)
		self.tb_rangos.setRowCount(totalfilas)
			
			
		fila =0
		acum=0
		for i in listarecuperada:
			self.tb_rangos.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_rangos.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_rangos.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
			self.tb_rangos.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
			self.tb_rangos.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
							
			fila=fila+1
			acum=acum+int(i[1])
		
		self.signal_total_rangos.setText(str(acum))
		
		
		
	def setearrango(self):
		inicio=int(self.signal_inicio_rg.text())
		final=int(self.signal_fin_rg.text())
		if self.rb_definir_seriea.isChecked():
			indice=1
		else:
			indice=2
		
		q.definirrango(inicio,final,indice)
		
		self.signal_inicio_rg.setText("")
		self.signal_fin_rg.setText("")
		
	def rangoselected(self):
		fila = self.tb_rangos.currentRow()
		inicio=self.tb_rangos.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		fin=self.tb_rangos.item(fila, 1).text()
		
		self.signal_inicio_rg.setText(str(inicio))
		self.signal_fin_rg.setText(str(fin))
		
	def bajaderango(self):
		inicio=int(self.signal_inicio_rg.text())
		
		if self.rb_definir_seriea.isChecked():
			indice=1
		else:
			indice=2
		
		q.cancelarrango(inicio)
		
		self.signal_inicio_rg.setText("")
		self.signal_fin_rg.setText("")
		
		
	def traerstock(self):
		
		listado= q.getstock()
		
		totalfilas=len(listado)
		self.tb_stock_rotulos.setRowCount(totalfilas)
		fila=0
		for i in listado:
			self.tb_stock_rotulos.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_stock_rotulos.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			fila = fila+1
			
	
	def listarasociados(self):
		
		
		
		
		listarecuperada=q.traerasociados()
		totalfilas=len(listarecuperada)
		self.tb_asociados_envios.setRowCount(totalfilas)
		
		self.tb_asociados_pedidos2.setRowCount(totalfilas)
		
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_asociados_envios.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			
			self.tb_asociados_pedidos2.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			
			
		
									
			fila=fila+1
		
		
	def asociado_selected(self):
					
			#ENVIOS
			
			fila = self.tb_asociados_envios.currentRow()
			asociado=str(self.tb_asociados_envios.item(fila, 0).text()) #SELECCIONO EL CONTENIDO DE LA FILA DE LA COLUMNA 0 SELECCIONADA
			self.signal_asociado_envios.setText(str(asociado))
			self.traerpedidos_agrupados(asociado)
			
			self.traerenvios()
			
	def envio_selected(self):
					
			#ENVIOS
			
			fila = self.tb_envioscreados.currentRow()
			id=str(self.tb_envioscreados.item(fila, 7).text()) #SELECCIONO EL CONTENIDO DE LA FILA DE LA COLUMNA 0 SELECCIONADA
			fecha_envio=str(self.tb_envioscreados.item(fila, 1).text())
			estado=str(self.tb_envioscreados.item(fila, 2).text())
			guia=str(self.tb_envioscreados.item(fila, 11).text())
			asociado=str(self.tb_envioscreados.item(fila, 0).text())
			fct=str(self.tb_envioscreados.item(fila, 12).text())
   
			self.signal_envio_fecha.setText(str(fecha_envio))
			self.signal_envio_estado.setText(str(estado))
			self.signal_envio_guia.setText(str(guia))
			self.signal_id_envio.setText(str(id))
			self.signal_envio_asociado.setText(str(asociado))
			self.signal_envio_fct.setText(str(fct))
			
			
			
			
	def asociado_selected_pedidos(self):
			#ESTAMPILLAS
			fila = self.tb_asociados_pedidos2.currentRow()
			asociado=self.tb_asociados_pedidos2.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA DE LA COLUMNA 0 SELECCIONADA
			registro=q.getrncyfs(str(asociado))[0]
			
			self.txt_rncyfs.setText(str(registro))
			
			
	def pedido_envios_selected(self):
		fila = self.tb_subpedidos_envios.currentRow()
		
		cantidad=self.tb_subpedidos_envios.item(fila, 0).text()
		especie=self.tb_subpedidos_envios.item(fila, 1).text() #SELECCIONO EL CONTENIDO DE LA FILA DE LA COLUMNA 0 SELECCIONADA
		fecha_emision=self.tb_subpedidos_envios.item(fila, 2).text()
		
		self.txt_cantidad_envios.setText(str(cantidad))
		self.txt_especie_envios.setText(str(especie))
		self.signal_emision.setText(str(fecha_emision))

	def eliminar_linea(self):
		if self.rb_estampillas.isChecked():
			fila = self.tb_estampillas.currentRow()
			inicio =int(self.tb_estampillas.item(fila, 2).text())
			r=c.cartel_opcion("ATENCION","DESEA ELIMINAR LA LINEA SELECCIONADA",2)
			
			if r==16384:
				q.eliminarLinea(inicio)
				c.cartel("ATENCION","LINEA ELIMINADA",1)

				q.corregirInicio(inicio,4)



			self.refresh_estampillas()

			
		else:
			
			fila = self.tb_estampillas.currentRow()
			inicio =int(self.tb_estampillas.item(fila, 2).text())
			r=c.cartel_opcion("ATENCION","DESEA ELIMINAR LA LINEA SELECCIONADA",2)
				
			if r==16384:
				q.eliminarLineaAnexo(inicio)
				c.cartel("ATENCION","LINEA ELIMINADA",1)

				q.corregirInicio(inicio,5)



			self.refresh_estampillas()

	
	
		
			
	def filtrar_asociados(self):
		if self.txt_asociados_envios:
			asociado= str("%"+self.txt_asociados_envios.text()+"%").upper()
		else:
			asociado="%"
		
		listarecuperada=q.traerasociadosFILTRO(asociado)
		totalfilas=len(listarecuperada)
		self.tb_asociados_envios.setRowCount(totalfilas)
			
		
		fila =0
		
		for i in listarecuperada:
			self.tb_asociados_envios.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
												
			fila=fila+1
	
	def filtrar_asociados_pedidos(self):
		if self.txt_buscar_asociados_pedidos:
			asociado= str("%"+self.txt_buscar_asociados_pedidos.text()+"%").upper()
		else:
			asociado="%"
		
		listarecuperada=q.traerasociadosFILTRO(asociado)
		totalfilas=len(listarecuperada)
		self.tb_asociados_pedidos2.setRowCount(totalfilas)
			
		
		fila =0
		
		for i in listarecuperada:
			self.tb_asociados_pedidos2.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
												
			fila=fila+1
		
			
		
		
		
		
				
		
	def traerpedidos_agrupados(self,asociado):

		if asociado:

			registro=q.traerregistro(str(asociado))
		else:
			registro='%'

		if self.rb_envios_iqr.isChecked():			
			
			tablarecuperada=q.subpedidosporfecha(registro)			

		elif self.rb_envios_estampillas.isChecked():
			

			tablarecuperada=q.estampillasPorFecha(registro)
			
		elif self.rb_envios_anexo.isChecked():
			tablarecuperada=q.anexosPorFecha(registro)



		totalfilas=len(tablarecuperada)
		self.tb_subpedidos_envios.setRowCount(totalfilas)
		fila=0
		for i in tablarecuperada:
			self.tb_subpedidos_envios.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
			self.tb_subpedidos_envios.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
			self.tb_subpedidos_envios.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				
			fila = fila+1

		
	def traerenvios(self):     
            
		try:
				
			desde=str(self.fecha_desde_envios.text())
			hasta=str(self.fecha_hasta_envios.text())
			if self.rb_facturados_envios.isChecked():
				estado='FACTURADO'
			elif self.rb_sin_facturar_envios.isChecked():
				estado='PREPARADO'
			elif self.rb_todos_envios.isChecked():
				estado='%'
			
			
			
			if self.cb_servicio_busqueda.currentText()=='':
				tipo='%'
			else:
				tipo=str(self.cb_servicio_busqueda.currentText())
				

			
			

			
			if self.rb_xasociado_envios.isChecked():
       
				fila = self.tb_asociados_envios.currentRow()	
				asociado=str(self.tb_asociados_envios.item(fila, 0).text())
				registro=q.traerregistro(str(asociado))
				
    
				if self.cbx_enviosxfecha.isChecked():
        
					
					tablarecuperada=q.getEnviosPorFecha(registro,desde,hasta,estado,tipo)
				
				else:
					tablarecuperada=q.getEnvios(registro,estado,tipo)
			else:
       
				if self.cbx_enviosxfecha.isChecked():
        
					tablarecuperada=q.getEnvios_ALLporFecha(desde,hasta,estado,tipo)
				
				else:
					tablarecuperada=q.getEnvios_ALL(estado,tipo)
				
				
			totalfilas=len(tablarecuperada)
			self.tb_envioscreados.setRowCount(totalfilas)
			fila=0
			for i in tablarecuperada:
				self.tb_envioscreados.setItem(fila,0,QtWidgets.QTableWidgetItem(str(i[0])))
				self.tb_envioscreados.setItem(fila,1,QtWidgets.QTableWidgetItem(str(i[1])))
				self.tb_envioscreados.setItem(fila,2,QtWidgets.QTableWidgetItem(str(i[2])))
				self.tb_envioscreados.setItem(fila,3,QtWidgets.QTableWidgetItem(str(i[3])))
				self.tb_envioscreados.setItem(fila,4,QtWidgets.QTableWidgetItem(str(i[4])))
				self.tb_envioscreados.setItem(fila,5,QtWidgets.QTableWidgetItem(str(i[5])))
				self.tb_envioscreados.setItem(fila,6,QtWidgets.QTableWidgetItem(str(i[6])))
				self.tb_envioscreados.setItem(fila,7,QtWidgets.QTableWidgetItem(str(i[7])))
				self.tb_envioscreados.setItem(fila,8,QtWidgets.QTableWidgetItem(str(i[8])))
				self.tb_envioscreados.setItem(fila,9,QtWidgets.QTableWidgetItem(str(i[9])))
				self.tb_envioscreados.setItem(fila,10,QtWidgets.QTableWidgetItem(str(i[10])))
				self.tb_envioscreados.setItem(fila,11,QtWidgets.QTableWidgetItem(str(i[11])))
				self.tb_envioscreados.setItem(fila,12,QtWidgets.QTableWidgetItem(str(i[12])))
				
				
				fila = fila+1
    
		except Exception as e:
			print(e)
      		
			
      	
	
		
		
		
			
		
		
		
		

		
def salir():
	exit()
		




		
class Pedido:
	
		def __init__(self,cantidad,rncyfs,num):
			self.cantidad=cantidad
			self.rncyfs=rncyfs
			self.numpedido=num
		
		        
		def asignar(self,inicio):
		
			self.inicio=inicio
			self.fin=inicio+self.cantidad-1
			self.disponibleinicio=inicio
			self.disponiblefin=inicio+self.cantidad-1
		
		def showrango(self):
			print ("Rango asignado:",self.inicio," - ",self.fin)		
		
		
		
def listarsub():
	for i in subpedidos:
		print ("pedido num: ",i.numpedido, "Cantidad Otorgada: ",i.cantidad,"Rango: ",i.inicio," - ",i.fin)		
		
	
		
		
class Subpedido():
	def __init__(self,inicio,fin,cantidad,numpedido,variedad,especie,camp,dav,categoria,registro):
		self.numpedido= numpedido
		self.cantidad=cantidad
		self.inicio=inicio
		self.fin=fin
		self.variedad=variedad
		self.especie= especie
		self.categoria= categoria
		self.camp= camp
		self.dav=dav
		self.registro=registro

def formatearfecha(fecha):
	pass
	'''fechaastring=str(fecha)

	dia=str(fechaastring[8:10])
	mes=str(fechaastring[5:7])
	year=str(fechaastring[0:4])

	fechacambiada=dia+"-"+mes+"-"+year
	
	return fechacambiada'''

		

def consultastock():
	print ("STOCK DISPONIBLE: ",(FINAL-INICIAL+1))


	
		
def actualizarangogeneral(cantidad):
	pass
	#Rango[0]=Rango[0]+cantidad
	INICIAL= INICIAL+cantidad
	
					
	


if __name__ == '__main__':
	q=bdquery()
	c=Mensaje()
	
	
	app = QApplication(sys.argv)
	
	MyWindow = VentanaPrincipal()
	MyWindow.llenarcombo()
	MyWindow.lockerdisponibles()
	MyWindow.listarasociados()
	MyWindow.traepedidos()
	MyWindow.listarlockers()
	MyWindow.refresh_estampillas()
	MyWindow.ver_estampillas()
	MyWindow.show()
	app.exec_()
	
	




		

   

    


